"""Investment Bot — Генерация Excel buy orders для STM-MCS.

ConversationHandler:
1. Объём ($) — ввод числа
2. Исключения — toggle кнопки
3. Мин. цена — ввод числа
4. Макс. цена — ввод числа
→ Генерация Excel → отправка файлом
"""

import json
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from telegram import (Update, InlineKeyboardButton,
                      InlineKeyboardMarkup)
from telegram.ext import (ContextTypes, ConversationHandler,
                          CallbackQueryHandler, MessageHandler,
                          filters)

log = logging.getLogger("invest")

# States
ST_VOLUME, ST_EXCLUDES, ST_MIN_PRICE, ST_MAX_PRICE, ST_DISCOUNT, ST_MIN_PROFIT, ST_NEW_KEY = range(7)

# Категории для исключения
CATEGORIES = [
    ("knives",    "🔪 Ножи/перчатки", ["★"]),
    ("cases",     "📦 Кейсы",         ["Case"]),
    ("capsules",  "💊 Капсулы",       ["Capsule"]),
    ("stickers",  "🏷️ Стикеры",      ["Sticker"]),
    ("graffiti",  "🎨 Граффити",      ["Graffiti", "Sealed Graffiti"]),
    ("music",     "🎵 Музыка",        ["Music Kit"]),
    ("patches",   "🎖️ Патчи",        ["Patch"]),
    ("souvenir",  "🔫 Сувенирное",    ["Souvenir"]),
    ("agents",    "📜 Агенты",        ["Agent"]),
    ("pins",      "🏅 Медали/пины",   ["Pin", "Medal"]),
]

APP_ID = 730
MARKET_FEE = 0.0  # без комиссии

SNIPER_DIR = os.path.expanduser(
    "~/.openclaw/agents/architect/projects/lis-sniper")
SNIPER_DB = os.path.join(SNIPER_DIR, "sniper.db")
OUTPUT_DIR = os.path.expanduser(
    "~/.openclaw/agents/architect/projects/investment-bot/buyorders")

# Ужесточённые пороги антибуста для buy orders
ANTIBOOST_TREND_THRESHOLD = -20.0    # med7d vs med30d
ANTIBOOST_VELOCITY_THRESHOLD = -5.0  # latest vs med7d (было -7%)
ANTIBOOST_MIN_SOLD_24H = 9           # минимум ликвидность
MIN_AGE_DAYS = 180                   # минимум 6 месяцев на рынке

# MarketCSGO API — мульти-ключи с round-robin ротацией
import threading as _th

_keys_lock = _th.Lock()
_mcsgo_keys: list[dict] = []  # [{key: str, alive: bool}]
_mcsgo_key_idx = 0
_mcsgo_req_counter = 0
_ROTATE_EVERY = 2  # переключаться каждые N запросов
_mcsgo_alert_fn = None  # callback для TG уведомлений


def _mask_key(key: str) -> str:
    """Первые 8 символов + ****."""
    return key[:8] + "****" if len(key) > 8 else key[:4] + "****"


def _load_keys():
    """Загрузить ключи из mcsgo_keys.txt, fallback на mcsgo_key.txt."""
    global _mcsgo_keys, _mcsgo_key_idx, _mcsgo_req_counter
    root = Path(__file__).parent.parent
    keys = []

    # Приоритет: mcsgo_keys.txt (мульти-ключи)
    multi_file = root / "mcsgo_keys.txt"
    if multi_file.exists():
        for line in multi_file.read_text().strip().splitlines():
            k = line.strip()
            if k and len(k) > 10 and not k.startswith("#"):
                keys.append(k)

    # Fallback: mcsgo_key.txt (один ключ)
    if not keys:
        single_file = root / "mcsgo_key.txt"
        if single_file.exists():
            k = single_file.read_text().strip()
            if k and len(k) > 10:
                keys.append(k)

    with _keys_lock:
        _mcsgo_keys = [{"key": k, "alive": True} for k in keys]
        _mcsgo_key_idx = 0
        _mcsgo_req_counter = 0

    if keys:
        log.info("🔑 MarketCSGO ключей загружено: %d (%s)",
                 len(keys), ", ".join(_mask_key(k) for k in keys))
    else:
        log.warning("⚠️ MarketCSGO: нет ключей!")


def _save_keys():
    """Сохранить текущие ключи в mcsgo_keys.txt."""
    root = Path(__file__).parent.parent
    with _keys_lock:
        lines = [kd["key"] for kd in _mcsgo_keys]
    (root / "mcsgo_keys.txt").write_text("\n".join(lines) + "\n")


def reload_mcsgo_keys():
    """Перечитать ключи из файла (после add/remove)."""
    _load_keys()


def get_mcsgo_keys_info() -> list[dict]:
    """Информация о ключах для TG UI [{masked, alive, current}]."""
    with _keys_lock:
        return [
            {
                "idx": i,
                "masked": _mask_key(kd["key"]),
                "alive": kd["alive"],
                "current": i == _mcsgo_key_idx,
            }
            for i, kd in enumerate(_mcsgo_keys)
        ]


def add_mcsgo_key(key: str) -> bool:
    """Добавить ключ. Возвращает False если дубликат."""
    global _mcsgo_key_idx, _mcsgo_req_counter
    with _keys_lock:
        if any(kd["key"] == key for kd in _mcsgo_keys):
            return False
        _mcsgo_keys.append({"key": key, "alive": True})
        # Удаляем мёртвые ключи (остаётся только живые + новый)
        _mcsgo_keys[:] = [kd for kd in _mcsgo_keys if kd["alive"] or kd["key"] == key]
        _mcsgo_key_idx = len(_mcsgo_keys) - 1  # переключаемся сразу на новый
        _mcsgo_req_counter = 0  # сбрасываем счётчик, чтобы начать с нового ключа
    _save_keys()
    log.info("🔑 MarketCSGO ключ добавлен: %s", _mask_key(key))
    return True


def remove_mcsgo_key(idx: int) -> bool:
    """Удалить ключ по индексу (0-based)."""
    global _mcsgo_key_idx
    with _keys_lock:
        if idx < 0 or idx >= len(_mcsgo_keys):
            return False
        removed = _mcsgo_keys.pop(idx)
        if _mcsgo_key_idx >= len(_mcsgo_keys):
            _mcsgo_key_idx = 0
    _save_keys()
    log.info("🗑 MarketCSGO ключ удалён: %s", _mask_key(removed["key"]))
    return True


def set_mcsgo_alert_fn(fn):
    """Установить callback для TG уведомлений при смерти ключа."""
    global _mcsgo_alert_fn
    _mcsgo_alert_fn = fn


def get_mcsgo_key() -> str:
    """Получить текущий ключ (round-robin каждые 2 запроса)."""
    global _mcsgo_key_idx, _mcsgo_req_counter
    with _keys_lock:
        if not _mcsgo_keys:
            return ""
        alive_keys = [i for i, kd in enumerate(_mcsgo_keys) if kd["alive"]]
        if not alive_keys:
            return _mcsgo_keys[0]["key"]  # все мертвы — вернуть первый для ошибки

        _mcsgo_req_counter += 1
        if _mcsgo_req_counter >= _ROTATE_EVERY:
            _mcsgo_req_counter = 0
            # Перейти на следующий живой ключ
            cur = _mcsgo_key_idx
            for _ in range(len(_mcsgo_keys)):
                cur = (cur + 1) % len(_mcsgo_keys)
                if _mcsgo_keys[cur]["alive"]:
                    _mcsgo_key_idx = cur
                    break

        return _mcsgo_keys[_mcsgo_key_idx]["key"]


def set_mcsgo_key(key: str):
    """Совместимость: установить единственный ключ."""
    global _mcsgo_keys, _mcsgo_key_idx
    with _keys_lock:
        _mcsgo_keys = [{"key": key, "alive": True}]
        _mcsgo_key_idx = 0
    _save_keys()
    log.info("🔑 MarketCSGO API ключ обновлён: %s...", key[:4])


def _on_bad_key(key: str):
    """Пометить ключ как мёртвый, переключиться на следующий."""
    global _mcsgo_key_idx
    with _keys_lock:
        for i, kd in enumerate(_mcsgo_keys):
            if kd["key"] == key:
                kd["alive"] = False
                log.error("💀 MarketCSGO ключ #%d мёртв: %s", i + 1, _mask_key(key))
                break

        # Переключиться на следующий живой
        alive = [i for i, kd in enumerate(_mcsgo_keys) if kd["alive"]]
        if alive:
            _mcsgo_key_idx = alive[0]
            log.info("🔄 Переключаюсь на ключ #%d: %s",
                     _mcsgo_key_idx + 1, _mask_key(_mcsgo_keys[_mcsgo_key_idx]["key"]))
        else:
            log.error("❌ ВСЕ MarketCSGO ключи мертвы!")

    # TG уведомление
    if _mcsgo_alert_fn:
        alive_count = sum(1 for kd in _mcsgo_keys if kd["alive"])
        total = len(_mcsgo_keys)
        dead_idx = next((i for i, kd in enumerate(_mcsgo_keys) if kd["key"] == key), 0)
        masked = _mask_key(key)
        if alive_count > 0:
            next_masked = _mask_key(_mcsgo_keys[_mcsgo_key_idx]["key"])
            _mcsgo_alert_fn(
                f"⚠️ Ключ #{dead_idx + 1} ({masked}) мёртв, "
                f"переключаюсь на #{_mcsgo_key_idx + 1} ({next_masked})\n"
                f"🔑 Активных: {alive_count}/{total}")
        else:
            _mcsgo_alert_fn(
                f"🚨 Ключ #{dead_idx + 1} ({masked}) мёртв — "
                f"ВСЕ {total} ключей мертвы!\n"
                f"Сканирование остановлено. Добавь новый через кнопку 🔑")


# При запуске — загружаем ключи
_load_keys()
MCSGO_BATCH_SIZE = 45                # макс предметов за запрос
MCSGO_RATE_LIMIT = 1.0               # 1 сек сон после каждого батча


def _excludes_kb(excluded: set) -> InlineKeyboardMarkup:
    """Клавиатура toggle-кнопок для исключений."""
    rows = []
    row = []
    for key, label, _ in CATEGORIES:
        mark = "❌" if key in excluded else "✅"
        row.append(InlineKeyboardButton(
            f"{mark} {label}", callback_data=f"bo:ex:{key}"))
        if len(row) == 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(
        "✅ Готово", callback_data="bo:ex:done")])
    return InlineKeyboardMarkup(rows)


def _get_steamwebapi_data() -> list:
    """Получить bulk данные из SteamWebAPI."""
    _src = os.path.join(SNIPER_DIR, "src")
    if _src not in sys.path:
        sys.path.insert(0, _src)
    saved_cwd = os.getcwd()
    os.chdir(SNIPER_DIR)
    try:
        import importlib
        # Загружаем .env lis-sniper для ключа
        env_path = os.path.join(SNIPER_DIR, ".env")
        if os.path.exists(env_path):
            with open(env_path) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        k, v = line.split("=", 1)
                        os.environ.setdefault(k.strip(), v.strip())
        
        import steamwebapi
        key = os.environ.get("STEAMWEBAPI_KEY", "")
        if key:
            steamwebapi.set_key(key)
        items = steamwebapi.get_bulk("cs2", max_items=50000)
        log.info("SteamWebAPI: %d предметов загружено", len(items))
        return items
    except Exception as e:
        log.error("SteamWebAPI error: %s", e)
        return []
    finally:
        os.chdir(saved_cwd)


# Кэш MarketCSGO bulk {name: price} — быстрая проверка наличия
_mcsgo_names_cache: set = set()
_mcsgo_names_ts: float = 0
_MCSGO_CACHE_TTL = 3600  # 1 час


_mcsgo_bulk_prices_cache: dict = {}  # {name: min_listing_price}

def _get_marketcsgo_names() -> set:
    """Быстрая загрузка множества названий через bulk (для фильтрации).
    Также кеширует min listing price в _mcsgo_bulk_prices_cache."""
    global _mcsgo_names_cache, _mcsgo_names_ts, _mcsgo_bulk_prices_cache
    
    if _mcsgo_names_cache and (time.time() - _mcsgo_names_ts < _MCSGO_CACHE_TTL):
        log.info("MarketCSGO names: из кэша (%d)", len(_mcsgo_names_cache))
        return _mcsgo_names_cache
    
    import httpx
    try:
        log.info("MarketCSGO: загрузка bulk для списка названий...")
        r = httpx.get(
            "https://market.csgo.com/api/v2/prices/class_instance/USD.json",
            timeout=60)
        data = r.json()
        if data.get("success"):
            names = set()
            prices = {}
            for val in data.get("items", {}).values():
                n = val.get("market_hash_name", "")
                if not n:
                    continue
                names.add(n)
                try:
                    p = float(val.get("price", 0) or 0)
                    pop = val.get("popularity_7d")
                    avg = float(val.get("avg_price", 0) or 0)
                    
                    if p <= 0:
                        continue
                    
                    # Приоритет: запись с реальными продажами
                    # (popularity > 0 и avg_price > 0)
                    has_sales = (pop is not None and int(pop) >= 3 
                                 and avg > 0)
                    
                    if n not in prices:
                        prices[n] = {"price": p, "avg": avg, 
                                     "has_sales": has_sales}
                    elif has_sales and not prices[n]["has_sales"]:
                        # Заменяем пустой стак на реальный
                        prices[n] = {"price": p, "avg": avg,
                                     "has_sales": True}
                    elif has_sales and prices[n]["has_sales"]:
                        # Оба реальные — берём с меньшей ценой
                        if p < prices[n]["price"]:
                            prices[n] = {"price": p, "avg": avg,
                                         "has_sales": True}
                except (ValueError, TypeError):
                    pass
            _mcsgo_names_cache = names
            _mcsgo_bulk_prices_cache = prices
            _mcsgo_names_ts = time.time()
            log.info("MarketCSGO: %d названий, %d цен загружено",
                     len(names), len(prices))
            return names
    except Exception as e:
        log.error("MarketCSGO bulk names error: %s", e)
    return _mcsgo_names_cache or set()


def _get_mcsgo_ref_prices(names: list) -> dict:
    """Получить негативные ref-цены через MarketCSGO API (batch по 50).
    
    Для каждого предмета:
      ref_price = min(median_7d, median_30d, latest)
    
    Returns: {name: ref_price}
    """
    import httpx
    import statistics
    
    now = time.time()
    result = {}
    checked = set()  # предметы которые API вернул (даже если отсеяны фильтром)
    batches = [names[i:i + MCSGO_BATCH_SIZE]
               for i in range(0, len(names), MCSGO_BATCH_SIZE)]
    
    key = get_mcsgo_key()
    log.info("MarketCSGO API: %d предметов, %d батчей (по %d), пауза %.1fс, key=%s...",
             len(names), len(batches), MCSGO_BATCH_SIZE, MCSGO_RATE_LIMIT, key[:4])
    
    for bi, batch in enumerate(batches):
        params = {"key": get_mcsgo_key()}
        for name in batch:
            params.setdefault("list_hash_name[]", [])
            if isinstance(params["list_hash_name[]"], str):
                params["list_hash_name[]"] = [params["list_hash_name[]"]]
        
        # httpx нужен список для repeated params
        url = "https://market.csgo.com/api/v2/get-list-items-info"
        query_parts = [f"key={get_mcsgo_key()}"]
        for name in batch:
            query_parts.append(f"list_hash_name[]={quote(name)}")
        full_url = url + "?" + "&".join(query_parts)
        
        try:
            # До 2 попыток на батч
            data = None
            for attempt in range(2):
                try:
                    t_before = time.time()
                    r = httpx.get(full_url, timeout=30)
                    t_after = time.time()
                    data = r.json()
                    break
                except Exception:
                    if attempt == 0:
                        time.sleep(MCSGO_RATE_LIMIT)
                        continue
                    raise
            if data is None:
                continue
            
            if not data.get("success"):
                error = data.get("error", "")
                log.warning("MarketCSGO batch %d/%d: success=false, error=%s (%.1fs)",
                           bi + 1, len(batches), error, t_after - t_before)
                if "Bad KEY" in str(error):
                    _on_bad_key(get_mcsgo_key())
                    alive = [kd for kd in _mcsgo_keys if kd["alive"]]
                    if not alive:
                        return {"error": "❌ Все MarketCSGO ключи мертвы!"}
                    # Retry этот батч с новым ключом
                    continue
                time.sleep(MCSGO_RATE_LIMIT)
                continue
            
            for name, info in data.get("data", {}).items():
                checked.add(name)
                history = info.get("history", [])
                if not history:
                    continue
                
                all_prices = []
                for entry in history:
                    try:
                        all_prices.append(float(entry[1]))
                    except (ValueError, TypeError, IndexError):
                        pass
                
                if not all_prices:
                    continue
                
                latest_price = all_prices[0]
                
                # --- Фильтр: стабильность цены (ослаблен) ---
                # stdev > 50% от медианы + >5 выбросов → исключить
                if len(all_prices) >= 10:
                    med_all = statistics.median(all_prices)
                    if med_all > 0:
                        outliers = sum(
                            1 for p in all_prices
                            if abs(p - med_all) / med_all > 0.50)
                        if outliers > 5:
                            stdev = statistics.stdev(all_prices)
                            if stdev / med_all > 0.50:
                                continue  # нестабильная цена
                
                # --- Фильтр: тренд на MCSGO (ослаблен) ---
                # avg последних 10 vs avg первых 10
                # Если дешевеет >25% → исключить
                if len(all_prices) >= 20:
                    avg_last10 = statistics.mean(all_prices[:10])
                    avg_first10 = statistics.mean(all_prices[-10:])
                    if avg_first10 > 0:
                        trend = (avg_last10 - avg_first10) / avg_first10 * 100
                        if trend < -25:
                            continue  # предмет дешевеет
                
                # Фильтруем по периодам
                prices_7d = []
                prices_30d = []
                for entry in history:
                    try:
                        ts = int(entry[0])
                        price = float(entry[1])
                    except (ValueError, TypeError, IndexError):
                        continue
                    age = now - ts
                    if age <= 7 * 86400:
                        prices_7d.append(price)
                    if age <= 30 * 86400:
                        prices_30d.append(price)
                
                med7 = statistics.median(prices_7d) if prices_7d else latest_price
                med30 = statistics.median(prices_30d) if prices_30d else latest_price
                
                # Проверка ликвидности: batch sales_7d >= 5 ИЛИ bulk has_sales
                sales_7d = 0
                for entry in history:
                    try:
                        if (now - int(entry[0])) <= 7 * 86400:
                            sales_7d += 1
                    except (ValueError, IndexError):
                        pass
                bulk_info = _mcsgo_bulk_prices_cache.get(name, {})
                bulk_has_sales = (isinstance(bulk_info, dict)
                                  and bulk_info.get("has_sales"))
                if sales_7d < 5 and not bulk_has_sales:
                    continue  # неликвид на MCSGO
                
                # Спайк: среднее из 3 последних продаж на MCSGO
                last3_avg = 0
                if len(all_prices) >= 3:
                    last3_avg = statistics.mean(all_prices[:3])
                
                # ref_price = min(med7d, bulk_avg, min_listing, last3_avg)
                # пессимистичный сценарий — берём наименьшее
                candidates = []
                if med7 > 0:
                    candidates.append(med7)
                if last3_avg > 0:
                    candidates.append(last3_avg)
                # med30 убрана — слишком пессимистично
                if isinstance(bulk_info, dict):
                    bulk_price = bulk_info.get("price", 0)
                    bulk_avg = bulk_info.get("avg", 0)
                    if bulk_avg > 0:
                        candidates.append(bulk_avg)
                    if bulk_price > 0 and bulk_info.get("has_sales"):
                        candidates.append(bulk_price)
                elif bulk_info:
                    candidates.append(float(bulk_info))
                
                ref_price = min(candidates) if candidates else 0
                if ref_price > 0:
                    result[name] = round(ref_price, 2)
                    
        except Exception as e:
            log.error("MarketCSGO batch %d/%d error: %s", bi + 1, len(batches), e)
        
        # Rate limit
        if bi < len(batches) - 1:
            time.sleep(MCSGO_RATE_LIMIT)
        
        # Прогресс каждые 10 батчей
        if (bi + 1) % 10 == 0:
            log.info("MarketCSGO API: %d/%d батчей (%d цен получено)",
                    bi + 1, len(batches), len(result))
    
    # Retry цикл — до 99% покрытия или пока улучшается
    max_retries = 5
    for retry_num in range(1, max_retries + 1):
        missing = [n for n in names if n not in result and n not in checked]
        total_processed = len(result) + len(checked)
        coverage = total_processed / len(names) * 100 if names else 100
        if coverage >= 99 or not missing:
            break
        
        log.info("MarketCSGO retry %d: %d предметов (покрытие %.0f%%), повтор...",
                 retry_num, len(missing), coverage)
        time.sleep(2)
        
        prev_count = len(result)
        r_batches = [missing[i:i + MCSGO_BATCH_SIZE]
                     for i in range(0, len(missing), MCSGO_BATCH_SIZE)]
        key_dead = False
        for bi, batch in enumerate(r_batches):
            url = "https://market.csgo.com/api/v2/get-list-items-info"
            query_parts = [f"key={get_mcsgo_key()}"]
            for name in batch:
                query_parts.append(f"list_hash_name[]={quote(name)}")
            full_url = url + "?" + "&".join(query_parts)
            try:
                resp = None
                for attempt in range(2):
                    try:
                        resp = httpx.get(full_url, timeout=30)
                        data = resp.json()
                        break
                    except Exception:
                        if attempt == 0:
                            time.sleep(MCSGO_RATE_LIMIT)
                            continue
                        raise
                if resp is None:
                    continue
                data = resp.json()
                if not data.get("success"):
                    if "Bad KEY" in str(data.get("error", "")):
                        _on_bad_key(get_mcsgo_key())
                        alive = [kd for kd in _mcsgo_keys if kd["alive"]]
                        if not alive:
                            key_dead = True
                            break
                        continue  # retry с новым ключом
                    time.sleep(MCSGO_RATE_LIMIT)
                    continue
                for name, info in data.get("data", {}).items():
                    checked.add(name)
                    history = info.get("history", [])
                    if not history:
                        continue
                    all_prices = [float(e[1]) for e in history if len(e) >= 2]
                    if not all_prices:
                        continue
                    now_ts = time.time()
                    # sales_7d фильтр убран в retry — ликвидность уже проверена в Фазе 1
                    prices_7d = [float(e[1]) for e in history
                                 if len(e) >= 2 and (now_ts - int(e[0])) <= 7*86400]
                    med7 = statistics.median(prices_7d) if prices_7d else all_prices[0]
                    last3_avg = statistics.mean(all_prices[:3]) if len(all_prices) >= 3 else 0
                    bulk_info = _mcsgo_bulk_prices_cache.get(name, {})
                    cands = []
                    if med7 > 0: cands.append(med7)
                    if last3_avg > 0: cands.append(last3_avg)
                    if isinstance(bulk_info, dict):
                        ba = bulk_info.get("avg", 0)
                        bp = bulk_info.get("price", 0)
                        if ba > 0: cands.append(ba)
                        if bp > 0 and bulk_info.get("has_sales"):
                            cands.append(bp)
                    ref = min(cands) if cands else 0
                    if ref > 0:
                        result[name] = round(ref, 2)
            except Exception as e:
                log.error("MarketCSGO retry %d batch error: %s", retry_num, e)
            if bi < len(r_batches) - 1:
                time.sleep(MCSGO_RATE_LIMIT)
        
        new_count = len(result)
        log.info("MarketCSGO retry %d: +%d цен, итого %d/%d (%.0f%%)",
                 retry_num, new_count - prev_count, new_count, len(names),
                 new_count / len(names) * 100)
        
        if key_dead:
            break
        if new_count == prev_count:
            log.info("MarketCSGO retry: нет улучшений, стоп")
            break
    
    log.info("MarketCSGO API: готово, %d/%d цен, %d checked, %d без ответа API (%.0f%% покрытие)",
             len(result), len(names), len(checked),
             len(names) - len(checked),
             len(checked) / len(names) * 100 if names else 0)
    return result


def _antiboost_check(item: dict) -> tuple[bool, list[str]]:
    """Ужесточённый антибуст для buy orders.
    
    Пороги строже чем в lis-sniper:
    - Тренд: -10% (было -20%)
    - Velocity: -5% (было -7%)
    - Ликвидность: 20 sold/24h (было 5)
    """
    import statistics
    
    med30 = item.get("pricemedian30d") or 0
    med7 = item.get("pricemedian7d") or 0
    latest_sell = item.get("pricelatestsell") or 0
    latest_list = item.get("pricelatest") or 0
    pmin = item.get("pricemin") or 0
    pmax = item.get("pricemax") or 0
    sold24h = item.get("sold24h") or 0
    unstable = item.get("unstable", False)
    sales = item.get("latest10steamsales") or item.get("latest10") or []
    
    reasons = []
    
    # 1. Ликвидность (ужесточена: 20 вместо 5)
    if sold24h < ANTIBOOST_MIN_SOLD_24H:
        reasons.append(f"Ликвидность: {sold24h}/{ANTIBOOST_MIN_SOLD_24H}")
    
    # 2. Дамп 30д (хайп убран)
    if med30 > 0 and latest_sell > 0:
        growth = (latest_sell - med30) / med30 * 100
        if growth < -20:
            reasons.append(f"Дамп: {growth:+.1f}%")
    
    # 3. Медиана vs текущий — убран (слишком строгий для buy orders)
    
    # 4. Волатильность — убрана для buy orders
    
    # 5. Скачок из latest10
    if sales and len(sales) >= 6:
        try:
            prices = [float(s[1]) for s in sales if len(s) >= 2]
            if len(prices) >= 6:
                avg_recent = statistics.mean(prices[:3])
                avg_older = statistics.mean(prices[3:])
                if avg_older > 0:
                    spike = (avg_recent - avg_older) / avg_older * 100
                    if abs(spike) > 15:
                        reasons.append(f"Скачок: {spike:+.1f}%")
        except (ValueError, TypeError):
            pass
    
    # 6. Тренд: med7d vs med30d (ужесточен: -10% вместо -20%)
    if med7 > 0 and med30 > 0:
        trend = (med7 - med30) / med30 * 100
        if trend < ANTIBOOST_TREND_THRESHOLD:
            reasons.append(f"Тренд: {trend:+.1f}%")
    
    # 7. Velocity: latest vs med7d (ужесточен: -5% вместо -7%)
    if med7 > 0 and latest_sell > 0:
        velocity = (latest_sell - med7) / med7 * 100
        if velocity < ANTIBOOST_VELOCITY_THRESHOLD:
            reasons.append(f"Velocity: {velocity:+.1f}%")
    
    # 8. Unstable
    if unstable:
        reasons.append("Unstable")
    
    passed = len(reasons) == 0
    return passed, reasons


def _is_too_young(item: dict) -> bool:
    """Проверить что предмет младше MIN_AGE_DAYS (6 месяцев)."""
    from datetime import datetime, timezone
    
    # firstseenat — ISO формат: "2018-08-22T23:00:00+00:00"
    first_seen = item.get("firstseenat", "")
    if not first_seen:
        return False  # нет данных — пропускаем фильтр
    
    try:
        if isinstance(first_seen, str):
            seen_dt = datetime.fromisoformat(first_seen.replace("Z", "+00:00"))
        else:
            return False
        
        age_days = (datetime.now(timezone.utc) - seen_dt).days
        return age_days < MIN_AGE_DAYS
    except Exception:
        return False


def _should_exclude(name: str, excluded: set) -> bool:
    """Проверить нужно ли исключить предмет."""
    for key, _, patterns in CATEGORIES:
        if key not in excluded:
            continue
        for pattern in patterns:
            if pattern in name:
                return True
    return False


def _build_items(raw_items: list, excluded: set,
                 min_price: float, max_price: float,
                 total_volume: float,
                 discount: float = 0,
                 min_profit: float = 0,
                 progress_cb=None) -> list:
    """Отобрать и отсортировать предметы с антибустом и проверкой MarketCSGO.
    
    Двухфазная фильтрация:
    1. Быстрая фильтрация (SteamWebAPI + bulk names) → кандидаты
    2. Batch запрос ref-цен MarketCSGO API → финальный расчёт маржи
    """
    # Фаза 1: быстрая фильтрация
    mcsgo_names = _get_marketcsgo_names()
    
    stats = {"total": 0, "no_buy": 0, "no_steam": 0, "low_sold": 0,
             "price_filter": 0, "excluded_cat": 0, "no_mcsgo": 0,
             "too_young": 0, "antiboost": 0, "no_margin": 0,
             "no_ref_price": 0, "passed": 0}
    
    candidates = []  # прошли быстрые фильтры

    for item in raw_items:
        name = item.get("markethashname", "")
        if not name:
            continue
        stats["total"] += 1

        buy_order = item.get("buyorderprice") or 0
        if buy_order <= 0:
            stats["no_buy"] += 1
            continue

        steam_price = item.get("pricemedian30d") or 0
        if steam_price <= 0:
            stats["no_steam"] += 1
            continue

        sold24h = item.get("sold24h") or 0
        if sold24h < ANTIBOOST_MIN_SOLD_24H:
            stats["low_sold"] += 1
            continue

        if buy_order < min_price or buy_order > max_price:
            stats["price_filter"] += 1
            continue

        if _should_exclude(name, excluded):
            stats["excluded_cat"] += 1
            continue

        if mcsgo_names and name not in mcsgo_names:
            stats["no_mcsgo"] += 1
            continue

        if _is_too_young(item):
            stats["too_young"] += 1
            continue

        ab_passed, _ = _antiboost_check(item)
        if not ab_passed:
            stats["antiboost"] += 1
            continue

        candidates.append({
            "name": name,
            "buy_order": buy_order,
            "steam_price": steam_price,
            "sold24h": sold24h,
            "item": item,
        })

    log.info("📊 Фаза 1: %d кандидатов из %d (нет buy=%d, нет steam=%d, "
             "sold=%d, цена=%d, кат=%d, MCS=%d, молод=%d, антибуст=%d)",
             len(candidates), stats["total"], stats["no_buy"],
             stats["no_steam"], stats["low_sold"], stats["price_filter"],
             stats["excluded_cat"], stats["no_mcsgo"], stats["too_young"],
             stats["antiboost"])

    if not candidates:
        return []

    # Фаза 2: получаем ref-цены MarketCSGO API (batch по 50)
    if progress_cb:
        progress_cb(f"⏳ Получаю цены MarketCSGO ({len(candidates)} предметов)...")
    
    candidate_names = [c["name"] for c in candidates]
    ref_prices = _get_mcsgo_ref_prices(candidate_names)
    
    # Проверка ошибки API ключа
    if isinstance(ref_prices, dict) and "error" in ref_prices:
        return ref_prices  # вернём ошибку наверх
    
    # Фаза 3: финальный расчёт маржи от ref-цен (с учётом скидки)
    discount_mult = 1 - (discount / 100) if discount > 0 else 1
    if discount > 0:
        log.info("📉 Скидка к buy order: %.0f%% (×%.2f)", discount, discount_mult)
    if min_profit > 0:
        log.info("📈 Фильтр мін. прибыли: %.0f%%", min_profit)
    
    results = []
    for c in candidates:
        name = c["name"]
        buy_price = round(c["buy_order"] * discount_mult, 2)
        
        ref_price = ref_prices.get(name)
        if not ref_price or ref_price <= 0:
            stats["no_ref_price"] += 1
            continue
        
        net = ref_price * (1 - MARKET_FEE)
        margin = ((net - buy_price) / buy_price * 100) if buy_price > 0 else 0
        if margin < min_profit or margin > 50:
            stats["no_margin"] += 1
            continue

        steam_url = (f"https://steamcommunity.com/market/listings/"
                     f"{APP_ID}/{quote(name)}")
        mcsgo_url = f"https://market.csgo.com/?s=&search={quote(name)}"

        stats["passed"] += 1
        results.append({
            "name": name,
            "buy_order": round(buy_price, 2),
            "steam_price": round(c["steam_price"], 2),
            "mcsgo_price": round(ref_price, 2),
            "net": round(net, 2),
            "margin": round(margin, 1),
            "volume": c["sold24h"],
            "url": steam_url,
            "mcsgo_url": mcsgo_url,
        })

    log.info("📊 Фаза 2: ref-цены=%d/%d, нет маржи=%d → итого=%d",
             len(ref_prices), len(candidates),
             stats["no_margin"], stats["passed"])
    
    # Сортировка по марже
    results.sort(key=lambda x: x["margin"], reverse=True)

    # Каждый предмет по 1 штуке, обрезаем по бюджету
    budget_left = total_volume
    for item in results:
        item["qty"] = 1
    
    # Обрезаем по бюджету
    if total_volume > 0:
        filtered = []
        for item in results:
            if budget_left >= item["buy_order"]:
                filtered.append(item)
                budget_left -= item["buy_order"]
        results = filtered

    log.info("Отобрано: %d предметов", len(results))
    return results


def _generate_excel(items: list, params: dict) -> str:
    """Создать Excel файл."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    today = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"buyorders_{today}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buy Orders"

    # Стили
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496",
                           fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                             fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                              fill_type="solid")

    # Информационная строка
    ws.merge_cells("A1:F1")
    info = (f"Объём: ${params['volume']:.2f} | "
            f"Цена: ${params['min_price']:.2f}-${params['max_price']:.2f} | "
            f"Предметов: {len(items)} | "
            f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws["A1"] = info
    ws["A1"].font = Font(bold=True, size=10)

    # Заголовки
    headers = ["№", "App ID", "Название", "Ссылка Steam",
               "Ссылка MarketCSGO", "Buy Order ($)", "Кол-во",
               "Маржа %", "Объём 24ч"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    # Данные
    total_cost = 0
    for i, item in enumerate(items, 1):
        row = i + 2
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=APP_ID).border = border
        ws.cell(row=row, column=3, value=item["name"]).border = border

        url_cell = ws.cell(row=row, column=4, value=item["url"])
        url_cell.hyperlink = item["url"]
        url_cell.font = Font(color="0563C1", underline="single")
        url_cell.border = border

        mcsgo_cell = ws.cell(row=row, column=5, value=item.get("mcsgo_url", ""))
        mcsgo_cell.hyperlink = item.get("mcsgo_url", "")
        mcsgo_cell.font = Font(color="0563C1", underline="single")
        mcsgo_cell.border = border

        ws.cell(row=row, column=6, value=item["buy_order"]).border = border
        ws.cell(row=row, column=7, value=item.get("qty", 1)).border = border

        margin_cell = ws.cell(row=row, column=8, value=item["margin"])
        margin_cell.border = border
        if item["margin"] >= 20:
            margin_cell.fill = green_fill
        elif item["margin"] >= 10:
            margin_cell.fill = yellow_fill

        ws.cell(row=row, column=9, value=item["volume"]).border = border

        total_cost += item["buy_order"] * item.get("qty", 1)

    # Итого
    total_row = len(items) + 3
    ws.cell(row=total_row, column=5, value="ИТОГО:").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=round(total_cost, 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=7,
            value=sum(it.get("qty", 1) for it in items)).font = Font(bold=True)

    # Ширина колонок
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10

    wb.save(filepath)
    log.info("Excel сохранён: %s (%d предметов, $%.2f)",
             filepath, len(items), total_cost)
    return filepath


# ============================================================
# ConversationHandler
# ============================================================

async def _cancel_and_menu(update: Update,
                           ctx: ContextTypes.DEFAULT_TYPE):
    """Выход из диалога при нажатии кнопки меню."""
    ctx.user_data.clear()
    # Передаём обработку в основной handle_text
    from tg_handlers import handle_text
    await handle_text(update, ctx)
    return ConversationHandler.END


_LAST_SETTINGS_FILE = Path(__file__).parent.parent / "bo_last_settings.json"


def _save_last_settings(settings: dict):
    try:
        import json
        _LAST_SETTINGS_FILE.write_text(json.dumps(settings))
    except Exception:
        pass


def _load_last_settings() -> dict | None:
    try:
        import json
        if _LAST_SETTINGS_FILE.exists():
            return json.loads(_LAST_SETTINGS_FILE.read_text())
    except Exception:
        pass
    return None


async def start_buyorders(update: Update,
                          ctx: ContextTypes.DEFAULT_TYPE):
    """Начало — спрашиваем объём или повторить."""
    ctx.user_data["bo_excludes"] = {
        "knives", "cases", "capsules", "stickers",
        "graffiti", "patches", "agents", "pins"
    }
    
    # Статус текущего ключа
    keys_info = get_mcsgo_keys_info()
    current = next((k for k in keys_info if k["current"]), None)
    if current:
        key_status = f"🔑 Ключ MCSGO: <code>{current['masked']}</code> {'✅' if current['alive'] else '💀'}"
    elif keys_info:
        key_status = f"🔑 Ключей: {sum(1 for k in keys_info if k['alive'])}/{len(keys_info)}"
    else:
        key_status = "🔑 Ключ MCSGO: ❌ не задан"

    last = _load_last_settings()
    kb = []
    if last:
        kb.append([InlineKeyboardButton(
            f"🔄 Повторить (${last.get('volume', '?')}, "
            f"${last.get('min_price', '?')}-${last.get('max_price', '?')}, "
            f"мін {last.get('min_profit', '?')}%)",
            callback_data="bo:repeat")])
    kb.append([InlineKeyboardButton("🔑 Сменить ключ", callback_data="bo:change_key")])

    text = (
        f"📊 <b>Создание БД STM-MCS</b>\n\n"
        f"{key_status}\n\n"
        f"Введи общий объём ордеров ($):\n"
        f"<i>Например: 100</i>"
    )
    if last:
        text += "\n\nИли повтори последние настройки:"

    await update.message.reply_text(
        text, parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(kb) if kb else None)
    return ST_VOLUME


async def got_volume(update: Update,
                     ctx: ContextTypes.DEFAULT_TYPE):
    """Получили объём → показываем исключения."""
    text = update.message.text.strip().replace("$", "").replace(",", ".")
    try:
        volume = float(text)
        if volume <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❌ Введи число > 0")
        return ST_VOLUME

    ctx.user_data["bo_volume"] = volume
    excluded = ctx.user_data.get("bo_excludes", set())

    await update.message.reply_text(
        f"✅ Объём: <b>${volume:.2f}</b>\n\n"
        f"Выбери категории для <b>исключения</b>:\n"
        f"(нажми чтобы переключить, потом «Готово»)",
        parse_mode="HTML",
        reply_markup=_excludes_kb(excluded))
    return ST_EXCLUDES


async def toggle_exclude(update: Update,
                         ctx: ContextTypes.DEFAULT_TYPE):
    """Toggle категории исключения."""
    q = update.callback_query
    await q.answer()
    data = q.data  # bo:ex:knives или bo:ex:done

    key = data.split(":")[-1]

    if key == "done":
        excluded = ctx.user_data.get("bo_excludes", set())
        excl_text = ", ".join(
            label for k, label, _ in CATEGORIES if k in excluded
        ) or "ничего"
        await q.message.edit_text(
            f"✅ Исключено: {excl_text}\n\n"
            f"Введи <b>минимальную</b> цену ордера ($):\n"
            f"<i>По умолчанию: 0.50</i>",
            parse_mode="HTML")
        return ST_MIN_PRICE

    excluded = ctx.user_data.get("bo_excludes", set())
    if key in excluded:
        excluded.discard(key)
    else:
        excluded.add(key)
    ctx.user_data["bo_excludes"] = excluded

    await q.message.edit_reply_markup(
        reply_markup=_excludes_kb(excluded))
    return ST_EXCLUDES


async def got_min_price(update: Update,
                        ctx: ContextTypes.DEFAULT_TYPE):
    """Получили мин. цену → спрашиваем макс."""
    text = update.message.text.strip().replace("$", "").replace(",", ".")
    if text in (".", "", "д", "ок"):
        min_price = 0.50  # дефолт
    else:
        try:
            min_price = float(text)
            if min_price < 0:
                raise ValueError
        except ValueError:
            await update.message.reply_text("❌ Введи число ≥ 0")
            return ST_MIN_PRICE

    ctx.user_data["bo_min_price"] = min_price

    await update.message.reply_text(
        f"✅ Мин. цена: <b>${min_price:.2f}</b>\n\n"
        f"Введи <b>максимальную</b> цену ордера ($):\n"
        f"<i>По умолчанию: 10.00</i>",
        parse_mode="HTML")
    return ST_MAX_PRICE


async def got_max_price(update: Update,
                        ctx: ContextTypes.DEFAULT_TYPE):
    """Получили макс. цену → спрашиваем скидку."""
    text = update.message.text.strip().replace("$", "").replace(",", ".")
    if text in (".", "", "д", "ок"):
        max_price = 10.0  # дефолт
    else:
        try:
            max_price = float(text)
            if max_price <= 0:
                raise ValueError
        except ValueError:
            await update.message.reply_text("❌ Введи число > 0")
            return ST_MAX_PRICE

    min_price = ctx.user_data.get("bo_min_price", 0)
    if max_price <= min_price:
        await update.message.reply_text(
            f"❌ Макс. цена должна быть > мин. (${min_price:.2f})")
        return ST_MAX_PRICE

    ctx.user_data["bo_max_price"] = max_price

    await update.message.reply_text(
        f"✅ Макс. цена: <b>${max_price:.2f}</b>\n\n"
        f"Скидка к buy order (<b>%</b>)?\n"
        f"На сколько % ниже текущего buy order ставить ордер.\n"
        f"<i>Например: 10 → ордер = buyorder × 0.90</i>\n"
        f"<i>0 = без скидки (по умолчанию)</i>",
        parse_mode="HTML")
    return ST_DISCOUNT


async def got_discount(update: Update,
                       ctx: ContextTypes.DEFAULT_TYPE):
    """Получили скидку → спрашиваем мин. прибыль."""
    text = update.message.text.strip().replace("%", "").replace(",", ".")
    if text in ("0", ".", "", "д", "ок", "нет", "-"):
        discount = 0.0
    else:
        try:
            discount = float(text)
            if discount < 0 or discount >= 100:
                await update.message.reply_text("❌ Введи число от 0 до 99")
                return ST_DISCOUNT
        except ValueError:
            await update.message.reply_text("❌ Введи число (% скидки)")
            return ST_DISCOUNT

    ctx.user_data["bo_discount"] = discount
    disc_text = f" (ордер = buyorder × {1 - discount/100:.2f})" if discount > 0 else ""

    await update.message.reply_text(
        f"✅ Скидка: <b>{discount:.0f}%</b>{disc_text}\n\n"
        f"Минимальный <b>% прибыли</b> при завозе?\n"
        f"<i>Например: 5 → только предметы с маржой ≥ 5%</i>\n"
        f"<i>По умолчанию: 8%</i>",
        parse_mode="HTML")
    return ST_MIN_PROFIT


async def got_min_profit(update: Update,
                         ctx: ContextTypes.DEFAULT_TYPE):
    """Получили мин. прибыль → генерируем Excel."""
    text = update.message.text.strip().replace("%", "").replace(",", ".")
    if text in (".", "", "д", "ок"):
        min_profit = 8.0  # дефолт 8%
    else:
        try:
            min_profit = float(text)
            if min_profit < 0:
                await update.message.reply_text("❌ Введи число ≥ 0")
                return ST_MIN_PROFIT
        except ValueError:
            await update.message.reply_text("❌ Введи число (% прибыли)")
            return ST_MIN_PROFIT

    volume = ctx.user_data.get("bo_volume", 100)
    excluded = ctx.user_data.get("bo_excludes", set())
    min_price = ctx.user_data.get("bo_min_price", 0)
    max_price = ctx.user_data.get("bo_max_price", 10)
    discount = ctx.user_data.get("bo_discount", 0)

    params = {
        "volume": volume,
        "min_price": min_price,
        "max_price": max_price,
        "discount": discount,
        "min_profit": min_profit,
        "excluded": list(excluded),
    }

    info_lines = []
    if discount > 0:
        info_lines.append(f"📉 Скидка: {discount:.0f}%")
    if min_profit > 0:
        info_lines.append(f"📈 Мин. прибыль: {min_profit:.0f}%")
    extra_text = "\n".join(info_lines)

    # Отправляем "генерирую..."
    msg = await update.message.reply_text(
        f"⏳ <b>Генерирую БД STM-MCS...</b>\n\n"
        f"💰 Объём: ${volume:.2f}\n"
        f"📊 Цена: ${min_price:.2f} — ${max_price:.2f}\n"
        f"{extra_text}\n"
        f"🚫 Исключено: {len(excluded)} категорий",
        parse_mode="HTML")

    try:
        # Загружаем данные
        raw = _get_steamwebapi_data()
        if not raw:
            await msg.edit_text("❌ Не удалось загрузить данные SteamWebAPI")
            ctx.user_data.clear()
            return ConversationHandler.END

        items = _build_items(raw, excluded, min_price, max_price,
                             volume, discount=discount, min_profit=min_profit)
        if isinstance(items, dict) and "error" in items:
            if "Bad KEY" in items["error"] or "мертвы" in items["error"]:
                await msg.edit_text(
                    f"❌ {items['error']}\n\n"
                    f"🔑 Добавь новый ключ или проверь существующие\n"
                    f"Отправь новый API ключ MarketCSGO (USD):")
                return ST_NEW_KEY
            await msg.edit_text(f"❌ {items['error']}")
            ctx.user_data.clear()
            return ConversationHandler.END
        if not items:
            await msg.edit_text("❌ Нет предметов по заданным параметрам")
            ctx.user_data.clear()
            return ConversationHandler.END

        filepath = _generate_excel(items, params)

        # Сохраняем настройки для повтора
        _save_last_settings({
            "volume": volume,
            "min_price": min_price,
            "max_price": max_price,
            "discount": discount,
            "min_profit": min_profit,
            "excluded": list(excluded),
        })

        total_cost = sum(it["buy_order"] * it.get("qty", 1) for it in items)
        avg_margin = sum(it["margin"] for it in items) / len(items)

        rl = []
        if discount > 0:
            rl.append(f"📉 Скидка: {discount:.0f}%")
        if min_profit > 0:
            rl.append(f"📈 Мин. прибыль: {min_profit:.0f}%")
        extra_r = "\n".join(rl) + "\n" if rl else ""

        await msg.edit_text(
            f"✅ <b>БД STM-MCS готова!</b>\n\n"
            f"📦 Предметов: {len(items)} (из {len(raw)} загруженных)\n"
            f"💰 Общая стоимость: ${total_cost:.2f}\n"
            f"📈 Средняя маржа: {avg_margin:.1f}%\n"
            f"📊 Топ маржа: {items[0]['margin']:.1f}% ({items[0]['name'][:30]})\n\n"
            f"{extra_r}"
            f"🛡 Антибуст: тренд ≥{ANTIBOOST_TREND_THRESHOLD}%, "
            f"velocity ≥{ANTIBOOST_VELOCITY_THRESHOLD}%\n"
            f"📈 Мин. ликвидность: {ANTIBOOST_MIN_SOLD_24H} sold/24ч\n"
            f"✅ Проверено на MarketCSGO",
            parse_mode="HTML")

        # Отправляем файл
        keys_info = get_mcsgo_keys_info()
        alive_keys = sum(1 for k in keys_info if k["alive"])
        keys_status = f"🔑 {alive_keys}/{len(keys_info)} ключей"

        with open(filepath, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=os.path.basename(filepath),
                caption=f"📊 Buy Orders | ${volume:.2f} | "
                        f"{len(items)} предметов"
                        f"{f' | мін.прибыль {min_profit:.0f}%' if min_profit > 0 else ''}"
                        f"\n{keys_status}",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🔑 Ключи MCSGO", callback_data="bo:keys")]
                ]))

    except Exception as e:
        log.error("Buyorders generation error: %s", e)
        await msg.edit_text(f"❌ Ошибка генерации: {e}")

    ctx.user_data.clear()
    return ConversationHandler.END


async def repeat_last(update: Update,
                      ctx: ContextTypes.DEFAULT_TYPE):
    """Повторить последние настройки — сразу генерация."""
    query = update.callback_query
    await query.answer()
    
    last = _load_last_settings()
    if not last:
        await query.edit_message_text("❌ Нет сохранённых настроек")
        return ConversationHandler.END
    
    volume = last.get("volume", 100)
    excluded = set(last.get("excluded", []))
    min_price = last.get("min_price", 0.5)
    max_price = last.get("max_price", 10)
    discount = last.get("discount", 0)
    min_profit = last.get("min_profit", 8)
    
    ctx.user_data["bo_volume"] = volume
    ctx.user_data["bo_excludes"] = excluded
    ctx.user_data["bo_min_price"] = min_price
    ctx.user_data["bo_max_price"] = max_price
    ctx.user_data["bo_discount"] = discount
    
    info_lines = []
    if discount > 0:
        info_lines.append(f"📉 Скидка: {discount:.0f}%")
    if min_profit > 0:
        info_lines.append(f"📈 Мин. прибыль: {min_profit:.0f}%")
    extra_text = "\n".join(info_lines)
    
    msg = await query.edit_message_text(
        f"⏳ <b>Генерирую БД STM-MCS (повтор)...</b>\n\n"
        f"💰 Объём: ${volume:.2f}\n"
        f"📊 Цена: ${min_price:.2f} — ${max_price:.2f}\n"
        f"{extra_text}\n"
        f"🚫 Исключено: {len(excluded)} категорий",
        parse_mode="HTML")
    
    try:
        raw = _get_steamwebapi_data()
        if not raw:
            await msg.edit_text("❌ Не удалось загрузить данные SteamWebAPI")
            ctx.user_data.clear()
            return ConversationHandler.END
        
        params = {
            "volume": volume, "min_price": min_price,
            "max_price": max_price, "discount": discount,
            "min_profit": min_profit, "excluded": list(excluded),
        }
        
        items = _build_items(raw, excluded, min_price, max_price,
                             volume, discount=discount, min_profit=min_profit)
        if isinstance(items, dict) and "error" in items:
            if "Bad KEY" in items["error"] or "мертвы" in items["error"]:
                await msg.edit_text(
                    f"❌ {items['error']}\n\n"
                    f"🔑 Добавь новый ключ или проверь существующие\n"
                    f"Отправь новый API ключ MarketCSGO (USD):")
                return ST_NEW_KEY
            await msg.edit_text(f"❌ {items['error']}")
            ctx.user_data.clear()
            return ConversationHandler.END
        if not items:
            await msg.edit_text("❌ Нет предметов по заданным параметрам")
            ctx.user_data.clear()
            return ConversationHandler.END
        
        filepath = _generate_excel(items, params)
        _save_last_settings(params)
        
        total_cost = sum(it["buy_order"] * it.get("qty", 1) for it in items)
        avg_margin = sum(it["margin"] for it in items) / len(items)
        
        caption = (
            f"📊 Buy Orders | ${volume:.2f} | {len(items)} предметов"
            f" | мин.прибыль {min_profit:.0f}%")
        
        await msg.delete()
        with open(filepath, "rb") as f:
            await update.effective_chat.send_document(f, caption=caption)
        
    except Exception as e:
        log.error("Repeat buyorders error: %s", e)
        await msg.edit_text(f"❌ Ошибка: {e}")
    
    ctx.user_data.clear()
    return ConversationHandler.END


async def cancel_buyorders(update: Update,
                           ctx: ContextTypes.DEFAULT_TYPE):
    """Отмена."""
    ctx.user_data.clear()
    await update.message.reply_text("❌ Генерация отменена")
    return ConversationHandler.END


async def change_key_cb(update: Update,
                        ctx: ContextTypes.DEFAULT_TYPE):
    """Callback bo:change_key — переход в ST_NEW_KEY из любого состояния."""
    q = update.callback_query
    await q.answer()
    keys_info = get_mcsgo_keys_info()
    alive = sum(1 for k in keys_info if k["alive"])
    await q.edit_message_text(
        f"🔑 <b>Смена ключа MarketCSGO</b>\n\n"
        f"Текущих ключей: {alive}/{len(keys_info)}\n\n"
        f"Отправь новый API ключ (USD):",
        parse_mode="HTML")
    return ST_NEW_KEY


async def got_new_key(update: Update,
                      ctx: ContextTypes.DEFAULT_TYPE):
    """Получили новый MarketCSGO API ключ — сохраняем и перезапускаем генерацию."""
    key = update.message.text.strip()
    if len(key) < 10 or " " in key:
        await update.message.reply_text("❌ Неверный формат ключа. Попробуй ещё:")
        return ST_NEW_KEY

    add_mcsgo_key(key)
    info = get_mcsgo_keys_info()
    alive = sum(1 for k in info if k["alive"])
    await update.message.reply_text(
        f"✅ Новый ключ применён: <code>{_mask_key(key)}</code>\n"
        f"🔑 Активных: {alive} ключ(ей)\n\n"
        f"⏳ Перезапускаю генерацию...",
        parse_mode="HTML")

    # Перезапускаем генерацию с сохранёнными параметрами
    return await got_min_profit(update, ctx)


# ============================================================
# Управление ключами MCSGO (inline callbacks)
# ============================================================

async def _keys_menu(update_or_query):
    """Показать меню ключей."""
    keys = get_mcsgo_keys_info()
    if not keys:
        text = "🔑 <b>MarketCSGO ключи</b>\n\nНет ключей. Добавь через кнопку ниже."
    else:
        alive = [k for k in keys if k["alive"]]
        dead = [k for k in keys if not k["alive"]]
        lines = [f"🔑 <b>MarketCSGO ключи</b> ({len(alive)}/{len(keys)} активных)\n"]
        for k in alive:
            if k["current"]:
                lines.append(f"▶️ #{k['idx']+1} <code>{k['masked']}</code> <b>(используется)</b>")
            else:
                lines.append(f"✅ #{k['idx']+1} <code>{k['masked']}</code>")
        if dead:
            lines.append(f"\n❌ Мёртвых: {len(dead)}")
        text = "\n".join(lines)

    btns = [[InlineKeyboardButton("➕ Добавить ключ", callback_data="bo:keys_add")]]
    if keys:
        btns.append([InlineKeyboardButton(f"🗑 Удалить ключ", callback_data="bo:keys_rm")])
    btns.append([InlineKeyboardButton("🔙 Закрыть", callback_data="bo:keys_close")])
    kb = InlineKeyboardMarkup(btns)

    if hasattr(update_or_query, "edit_message_text"):
        await update_or_query.edit_message_text(text, parse_mode="HTML", reply_markup=kb)
    else:
        await update_or_query.reply_text(text, parse_mode="HTML", reply_markup=kb)


async def cb_keys(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Callback: bo:keys — показать меню."""
    q = update.callback_query
    await q.answer()
    await _keys_menu(q)


async def cb_keys_add(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Callback: bo:keys_add — запросить ключ."""
    q = update.callback_query
    await q.answer()
    await q.edit_message_text(
        "🔑 Отправь новый API ключ MarketCSGO (USD).\n"
        "Ключ будет добавлен к существующим.\n\n"
        "/cancel для отмены")
    ctx.user_data["_awaiting_mcsgo_key"] = True


async def cb_keys_rm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Callback: bo:keys_rm — показать кнопки удаления."""
    q = update.callback_query
    await q.answer()
    keys = get_mcsgo_keys_info()
    if not keys:
        await q.edit_message_text("Нет ключей для удаления.")
        return
    btns = []
    for i, k in enumerate(keys):
        status = "✅" if k["alive"] else "💀"
        btns.append([InlineKeyboardButton(
            f"🗑 #{i+1} {status} {k['masked']}",
            callback_data=f"bo:keys_rm:{i}")])
    btns.append([InlineKeyboardButton("🔙 Назад", callback_data="bo:keys")])
    await q.edit_message_text(
        "🗑 Выбери ключ для удаления:", reply_markup=InlineKeyboardMarkup(btns))


async def cb_keys_rm_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Callback: bo:keys_rm:N — удалить ключ."""
    q = update.callback_query
    await q.answer()
    idx = int(q.data.split(":")[-1])
    info = get_mcsgo_keys_info()
    if idx < 0 or idx >= len(info):
        await q.edit_message_text("❌ Неверный номер")
        return
    remove_mcsgo_key(idx)
    await _keys_menu(q)


async def cb_keys_close(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Callback: bo:keys_close — закрыть."""
    q = update.callback_query
    await q.answer()
    await q.edit_message_text("🔑 Меню ключей закрыто.")


async def on_mcsgo_key_message(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Обработчик текстового сообщения с ключом (после bo:keys_add)."""
    if not ctx.user_data.get("_awaiting_mcsgo_key"):
        return  # не ждём ключ
    ctx.user_data.pop("_awaiting_mcsgo_key", None)
    key = update.message.text.strip()
    if len(key) < 10:
        await update.message.reply_text("❌ Ключ слишком короткий")
        return
    ok = add_mcsgo_key(key)
    if ok:
        info = get_mcsgo_keys_info()
        await update.message.reply_text(
            f"✅ Ключ добавлен: <code>{_mask_key(key)}</code>\n"
            f"🔑 Всего ключей: {len(info)}",
            parse_mode="HTML")
    else:
        await update.message.reply_text("❌ Ключ уже существует")


def get_keys_handlers() -> list:
    """Handlers для inline-кнопок управления ключами MCSGO."""
    return [
        CallbackQueryHandler(cb_keys, pattern=r"^bo:keys$"),
        CallbackQueryHandler(cb_keys_add, pattern=r"^bo:keys_add$"),
        CallbackQueryHandler(cb_keys_rm, pattern=r"^bo:keys_rm$"),
        CallbackQueryHandler(cb_keys_rm_confirm, pattern=r"^bo:keys_rm:\d+$"),
        CallbackQueryHandler(cb_keys_close, pattern=r"^bo:keys_close$"),
    ]


def get_conversation_handler() -> ConversationHandler:
    """Создать ConversationHandler для buy orders."""
    _MENU_RE = r"^(📊 Инвестиции|🔄 Круги|🌐 Прокси|⚙️ Настройки|📜 История|📊 Создать БД STM-MCS|/start|/cancel)$"
    _NOT_MENU = filters.TEXT & ~filters.COMMAND & ~filters.Regex(_MENU_RE)
    return ConversationHandler(
        entry_points=[
            MessageHandler(
                filters.Regex(r"^📊 Создать БД STM-MCS$"),
                start_buyorders),
        ],
        states={
            ST_VOLUME: [
                CallbackQueryHandler(repeat_last, pattern=r"^bo:repeat$"),
                CallbackQueryHandler(change_key_cb, pattern=r"^bo:change_key$"),
                MessageHandler(_NOT_MENU, got_volume),
            ],
            ST_EXCLUDES: [
                CallbackQueryHandler(toggle_exclude,
                                     pattern=r"^bo:ex:"),
                CallbackQueryHandler(change_key_cb, pattern=r"^bo:change_key$"),
            ],
            ST_MIN_PRICE: [
                CallbackQueryHandler(change_key_cb, pattern=r"^bo:change_key$"),
                MessageHandler(_NOT_MENU, got_min_price),
            ],
            ST_MAX_PRICE: [
                CallbackQueryHandler(change_key_cb, pattern=r"^bo:change_key$"),
                MessageHandler(_NOT_MENU, got_max_price),
            ],
            ST_DISCOUNT: [
                CallbackQueryHandler(change_key_cb, pattern=r"^bo:change_key$"),
                MessageHandler(_NOT_MENU, got_discount),
            ],
            ST_MIN_PROFIT: [
                CallbackQueryHandler(change_key_cb, pattern=r"^bo:change_key$"),
                MessageHandler(_NOT_MENU, got_min_profit),
            ],
            ST_NEW_KEY: [
                MessageHandler(_NOT_MENU, got_new_key),
            ],
        },
        fallbacks=[
            MessageHandler(filters.Regex(r"^📊 Создать БД STM-MCS$"),
                           start_buyorders),
            MessageHandler(
                filters.Regex(r"^(📊 Инвестиции|🔄 Круги|🌐 Прокси|⚙️ Настройки|📜 История)$"),
                _cancel_and_menu),
            MessageHandler(filters.Regex(r"^(/cancel|/start)$"),
                           cancel_buyorders),
        ],
        per_message=False,
        allow_reentry=True,
    )
